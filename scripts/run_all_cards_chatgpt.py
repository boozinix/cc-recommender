#!/usr/bin/env python3
"""
Run ChatGPT summaries for every card in public/cards.csv and write
card_name, summary to a new CSV (and optionally Excel).

Requires: OPENAI_API_KEY in environment.
Usage:
  python run_all_cards_chatgpt.py
  python run_all_cards_chatgpt.py --output summaries.xlsx   # Excel instead of CSV
"""

import argparse
import csv
import os
import sys
import time
from pathlib import Path

# Scripts live in scripts/, cards live in public/
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
CARDS_CSV = REPO_ROOT / "public" / "cards.csv"
DEFAULT_OUTPUT_CSV = SCRIPT_DIR / "chatgpt_card_summaries.csv"

# Allow importing get_card_summary when run from repo root (e.g. python scripts/run_all_cards_chatgpt.py)
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))


def get_card_names_from_csv(path: Path) -> list[str]:
    """Read card_name (first column) from cards.csv."""
    names = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if "card_name" not in (reader.fieldnames or []):
            raise ValueError("cards.csv must have a 'card_name' column")
        for row in reader:
            name = (row.get("card_name") or "").strip()
            if name:
                names.append(name)
    return names


def main():
    parser = argparse.ArgumentParser(description="Run ChatGPT summary for every card in cards.csv")
    parser.add_argument(
        "--output", "-o",
        default=str(DEFAULT_OUTPUT_CSV),
        help="Output file path (default: scripts/chatgpt_card_summaries.csv)",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.5,
        help="Seconds to wait between API calls (default: 0.5)",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Skip cards that already have a row in the output file",
    )
    args = parser.parse_args()

    if not CARDS_CSV.exists():
        print(f"Error: Cards file not found: {CARDS_CSV}", file=sys.stderr)
        sys.exit(1)

    try:
        from get_card_summary import get_card_summary
    except ImportError as e:
        print("Error: Could not import get_card_summary.", file=sys.stderr)
        if "openai" in str(e).lower():
            print("Install the OpenAI package: pip install openai   (or pip3 install openai)", file=sys.stderr)
        else:
            print("Run this script from the repo root: python scripts/run_all_cards_chatgpt.py", file=sys.stderr)
        sys.exit(1)

    if not os.environ.get("OPENAI_API_KEY", "").strip():
        print(
            "Error: OPENAI_API_KEY is not set.\n"
            "Set it in this terminal, then run again:\n"
            "  export OPENAI_API_KEY=\"sk-your-key-here\"\n"
            "Get a key at https://platform.openai.com/api-keys",
            file=sys.stderr,
        )
        sys.exit(1)

    card_names = get_card_names_from_csv(CARDS_CSV)
    print(f"Found {len(card_names)} cards in {CARDS_CSV}")

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Resume: load existing card_name -> summary so we skip those
    existing: dict[str, str] = {}
    is_excel = output_path.suffix.lower() in (".xlsx", ".xls")
    if args.resume and output_path.exists():
        if is_excel:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(output_path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 2 and row[0]:
                        existing[str(row[0]).strip()] = str(row[1] or "").strip()
                wb.close()
            except Exception as e:
                print(f"Could not read existing Excel: {e}", file=sys.stderr)
        else:
            with open(output_path, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    name = (row.get("card_name") or "").strip()
                    if name:
                        existing[name] = row.get("summary", "")
        if existing:
            print(f"Resume: {len(existing)} cards already in output, skipping those.")

    if is_excel:
        try:
            import openpyxl
        except ImportError:
            print("For Excel output install openpyxl: pip install openpyxl", file=sys.stderr)
            sys.exit(1)

    # Fetch summaries (skip already-done when --resume)
    new_rows: list[tuple[str, str]] = []
    for i, card_name in enumerate(card_names, 1):
        if args.resume and card_name in existing:
            continue
        print(f"[{i}/{len(card_names)}] {card_name} ...")
        try:
            summary = get_card_summary(card_name)
            new_rows.append((card_name, summary))
        except Exception as e:
            print(f"  Error: {e}", file=sys.stderr)
            new_rows.append((card_name, f"[Error: {e}]"))
        if args.delay > 0:
            time.sleep(args.delay)

    # Merge existing + new, then output in original card order
    all_results = dict(existing)
    for name, summary in new_rows:
        all_results[name] = summary
    rows_to_write = [(name, all_results[name]) for name in card_names if name in all_results]

    if is_excel:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summaries"
        ws.append(["card_name", "summary"])
        for name, summary in rows_to_write:
            ws.append([name, summary])
        wb.save(output_path)
        print(f"Wrote {len(rows_to_write)} rows to {output_path}")
    else:
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["card_name", "summary"])
            writer.writerows(rows_to_write)
        print(f"Wrote {len(rows_to_write)} rows to {output_path}")


if __name__ == "__main__":
    main()
